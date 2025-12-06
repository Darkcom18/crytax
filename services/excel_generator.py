"""
Excel Report Generator
Generates Excel reports for tax calculations
"""

from typing import List, Dict
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from models.transaction import Transaction


class ExcelGenerator:
    """Generate Excel tax reports"""

    def __init__(self):
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_tax_report(
        self,
        transactions: List[Transaction],
        output_path: str,
        personal_info: Dict = None,
        tax_summary: Dict = None,
        tax_details: pd.DataFrame = None,
    ):
        """
        Generate comprehensive Excel tax report.

        Args:
            transactions: List of transactions
            output_path: Path to save Excel file
            personal_info: Personal information dict
            tax_summary: Tax summary data
            tax_details: DataFrame with detailed tax calculations
        """
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create sheets
        self._create_summary_sheet(wb, personal_info, tax_summary)
        self._create_transactions_sheet(wb, transactions)

        if tax_details is not None and not tax_details.empty:
            self._create_tax_details_sheet(wb, tax_details)

        # Save workbook
        wb.save(output_path)

    def _create_summary_sheet(
        self, wb: Workbook, personal_info: Dict, tax_summary: Dict
    ):
        """Create summary sheet with personal info and tax summary"""
        ws = wb.create_sheet("Tổng quan thuế")

        # Title
        ws["A1"] = "BÁO CÁO THUẾ GIAO DỊCH TIỀN ĐIỆN TỬ"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Date
        ws["A2"] = f"Ngày tạo: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells("A2:D2")
        ws["A2"].alignment = Alignment(horizontal="center")

        row = 4

        # Personal Information
        if personal_info:
            ws[f"A{row}"] = "THÔNG TIN CÁ NHÂN"
            ws[f"A{row}"].font = self.title_font
            row += 1

            info_data = [
                ["Họ và tên:", personal_info.get("name", "N/A")],
                ["CMND/CCCD:", personal_info.get("id_number", "N/A")],
                ["Địa chỉ:", personal_info.get("address", "N/A")],
                ["Số điện thoại:", personal_info.get("phone", "N/A")],
            ]

            for label, value in info_data:
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = value
                row += 1

            row += 1

        # Tax Summary
        if tax_summary:
            ws[f"A{row}"] = "TỔNG QUAN THUẾ"
            ws[f"A{row}"].font = self.title_font
            row += 1

            # Header
            headers = ["Hạng mục", "Giá trị (VND)"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = self.border

            row += 1

            # Tax data
            tax_data = [
                ["Tổng số giao dịch", tax_summary.get("total_transactions", 0)],
                ["Thuế chuyển nhượng (0.1%)", tax_summary.get("total_transfer_tax", 0)],
                [
                    "Thuế thu nhập khác (10%)",
                    tax_summary.get("total_other_income_tax", 0),
                ],
                ["Tổng thuế phải nộp", tax_summary.get("total_tax", 0)],
                ["Tổng lãi/lỗ", tax_summary.get("total_profit_loss", 0)],
            ]

            for label, value in tax_data:
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"A{row}"].border = self.border

                cell = ws[f"B{row}"]
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = "#,##0"
                else:
                    cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal="right")
                row += 1

            # Tax by token (if available)
            by_token = tax_summary.get("by_token", {})
            if by_token:
                row += 1
                ws[f"A{row}"] = "THUẾ THEO TOKEN"
                ws[f"A{row}"].font = self.title_font
                row += 1

                # Header
                for col, header in enumerate(["Token", "Thuế (VND)"], start=1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = self.border
                row += 1

                for token, amount in sorted(
                    by_token.items(), key=lambda x: x[1], reverse=True
                ):
                    ws[f"A{row}"] = token
                    ws[f"A{row}"].border = self.border

                    cell = ws[f"B{row}"]
                    cell.value = amount
                    cell.number_format = "#,##0"
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal="right")
                    row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

    def _create_transactions_sheet(self, wb: Workbook, transactions: List[Transaction]):
        """Create transactions sheet"""
        ws = wb.create_sheet("Chi tiết giao dịch")

        # Convert to DataFrame
        df = pd.DataFrame([tx.to_dict() for tx in transactions])

        # Select and rename columns
        display_columns = {
            "date": "Ngày",
            "type": "Loại",
            "token": "Token",
            "amount": "Số lượng",
            "value_vnd": "Giá trị (VND)",
            "source": "Nguồn",
            "chain": "Chain",
            "tx_hash": "TX Hash",
        }

        df = df[list(display_columns.keys())]
        df.columns = list(display_columns.values())

        # Format date
        df["Ngày"] = pd.to_datetime(df["Ngày"]).dt.strftime("%d/%m/%Y %H:%M")

        # Write headers
        for col, header in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # Write data
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), start=2
        ):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.border

                # Format number columns
                if c_idx in [4, 5]:  # Số lượng, Giá trị
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _create_tax_details_sheet(self, wb: Workbook, tax_details: pd.DataFrame):
        """Create tax details sheet"""
        ws = wb.create_sheet("Chi tiết tính thuế")

        # Prepare data
        df = tax_details.copy()

        # Extract transaction info
        df["Ngày"] = df["transaction"].apply(
            lambda x: x.date.strftime("%d/%m/%Y %H:%M")
        )
        df["Token"] = df["transaction"].apply(lambda x: x.token)
        df["Loại"] = df["transaction"].apply(lambda x: x.type.value)
        df["Số lượng"] = df["transaction"].apply(lambda x: x.amount)

        # Select and rename columns
        display_df = df[
            [
                "Ngày",
                "Token",
                "Loại",
                "Số lượng",
                "cost_basis",
                "profit_loss",
                "tax_amount",
                "tax_type",
            ]
        ].copy()

        display_df.columns = [
            "Ngày",
            "Token",
            "Loại",
            "Số lượng",
            "Giá vốn (VND)",
            "Lãi/Lỗ (VND)",
            "Thuế (VND)",
            "Loại thuế",
        ]

        # Write headers
        for col, header in enumerate(display_df.columns, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # Write data
        for r_idx, row in enumerate(
            dataframe_to_rows(display_df, index=False, header=False), start=2
        ):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.border

                # Format number columns
                if c_idx in [4, 5, 6, 7]:  # Number columns
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        # Adjust column widths
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 20
